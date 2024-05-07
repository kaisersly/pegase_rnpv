from collections import defaultdict
from copy import copy
from dataclasses import dataclass
import re
import openpyxl
from openpyxl.styles.cell_style import StyleArray
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from parse_pv import parse_pv, PV


def merge_pvs(
    merge_file: str,
    merge_pv_s: list[PV],
    template_file: str,
):
    merge = load_template(template_file)
    for merge_pv in merge_pv_s:
        add_pv(merge, merge_pv)
    save_merge(merge, merge_file)
    return merge_file


def merge_pv(merge_file: str, merge_pv: PV, template_file: str):
    return merge_pvs(merge_file, [merge_pv], template_file)


@dataclass
class Merge:
    wb: Workbook
    template_ws: Worksheet
    coordinates: dict[str, list[tuple[int, int]]]
    even_style: list[StyleArray]
    even_height: int
    odd_style: list[StyleArray]
    odd_height: int


required_merge_fields = ["CODE APPRENANT"]


def get_coordinates(template_ws: Worksheet) -> dict[str, list[tuple[int, int]]]:
    merge_field_regexp = r"#(.+)#"
    coordinates = defaultdict(list)
    # Récupérer les coordonnées des champs de fusion de la forme #FIELD#
    for row in template_ws.rows:
        for cell in row:
            value = cell.value or ""
            x = re.search(merge_field_regexp, value)
            if x:
                merge_field = x.group(1)
                coordinates[merge_field].append((cell.column, cell.row))
    # Vérifier que les champs de fusion requis sont présents
    absent_required_merge_fields = []
    for req in required_merge_fields:
        if req not in coordinates:
            absent_required_merge_fields.append(req)
    if absent_required_merge_fields:
        absent_fields_str = ", ".join([f"#{x}#" for x in absent_required_merge_fields])
        error = f"Champ(s) de fusion {absent_fields_str} manquant"
        raise Exception(error)
    # Vérifier que les champs de fusion de l'étudiant sont sur la même ligne
    etudiant_row_number = None
    for merge_field, field_coordinates in coordinates.items():
        for column, row in field_coordinates:
            if merge_field in ["CODE APPRENANT", "NOM", "PRENOM"]:
                if etudiant_row_number is None:
                    etudiant_row_number = row
                else:
                    if row != etudiant_row_number:
                        raise Exception(
                            "Les champs CODE APPRENANT, NOM et PRENOM doivent se trouver sur la même ligne"
                        )
    return coordinates


def get_etudiant_row_number(coordinates: dict[str, tuple[int, int]]) -> int:
    etudiant_row_number = coordinates["CODE APPRENANT"][0][1]
    return etudiant_row_number


def load_styles(
    template_ws: Worksheet, etudiant_row_number: int
) -> tuple[list[StyleArray], int, list[StyleArray], int]:
    even_row = next(
        template_ws.iter_rows(min_row=etudiant_row_number, max_row=etudiant_row_number)
    )
    even_style = [copy(cell._style) for cell in even_row]
    even_height = template_ws.row_dimensions[etudiant_row_number].height
    odd_row = next(
        template_ws.iter_rows(
            min_row=etudiant_row_number + 1, max_row=etudiant_row_number + 1
        )
    )
    odd_style = [copy(cell._style) for cell in odd_row]
    odd_height = template_ws.row_dimensions[etudiant_row_number + 1].height
    return even_style, even_height, odd_style, odd_height


def load_template(template_file: str) -> Merge:
    wb = openpyxl.load_workbook(filename=template_file)
    template_ws = wb.worksheets[0]
    template_ws.title = "TEMPLATE"
    coordinates = get_coordinates(template_ws)
    etudiant_row_number = get_etudiant_row_number(coordinates)
    even_style, even_height, odd_style, odd_height = load_styles(
        template_ws, etudiant_row_number
    )
    return Merge(
        wb=wb,
        template_ws=template_ws,
        coordinates=coordinates,
        even_style=even_style,
        even_height=even_height,
        odd_style=odd_style,
        odd_height=odd_height,
    )


def add_fixed_merge_fields(ws: Worksheet, fixed_merge_fields, coordinates):
    for merge_field, value in fixed_merge_fields.items():
        if merge_field in coordinates:
            for field_coordinates in coordinates[merge_field]:
                column_number, row_number = field_coordinates
                cell = ws.cell(row_number, column_number)
                cell.value = cell.value.replace(f"#{merge_field}#", value)


def add_etudiant(
    pv_ws: Worksheet,
    merge: Merge,
    first_etudiant_row_number: int,
    i: int,
    etudiant: dict[str, str],
):
    coordinates = merge.coordinates
    row_number = first_etudiant_row_number + i
    pv_ws.insert_rows(row_number)
    # Données
    for merge_field in etudiant.keys():
        if merge_field in coordinates:
            field_coordinates = coordinates[merge_field]
            for column_number, _ in field_coordinates:
                cell = pv_ws.cell(row_number, column_number)
                cell.value = etudiant[merge_field]
    # Style
    if i % 2 == 0:
        row_style = merge.even_style
        row_height = merge.even_height
    else:
        row_style = merge.odd_style
        row_height = merge.odd_height
    pv_ws.row_dimensions[row_number].height = row_height
    row = next(pv_ws.iter_rows(min_row=row_number, max_row=row_number))
    for i, cell in enumerate(row):
        cell._style = row_style[i]


def add_pv(merge: Merge, merge_pv: PV):
    etudiants = merge_pv.etudiants
    pv_ws = merge.wb.copy_worksheet(merge.template_ws)
    pv_ws.title = merge_pv.fixed_merge_fields["FORMATION_CODE"]
    add_fixed_merge_fields(pv_ws, merge_pv.fixed_merge_fields, merge.coordinates)
    first_etudiant_row_number = merge.coordinates["CODE APPRENANT"][0][1]
    pv_ws.delete_rows(first_etudiant_row_number)
    for i, etudiant in enumerate(etudiants):
        add_etudiant(pv_ws, merge, first_etudiant_row_number, i, etudiant)


def save_merge(merge: Merge, path: str):
    # Supprimer la feuille template
    merge.wb.remove(merge.template_ws)
    merge.wb.save(path)


if __name__ == "__main__":
    from time import time

    template_file = "template-pv.xlsx"
    merge_file = f"/mnt/c/Temp/PV-{int(time())}.xlsx"
    pv = parse_pv("pv-de-jury.csv")

    merge_pv(merge_file, pv, template_file)
